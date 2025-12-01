"""
Unified pipeline: 10-K HTML → AutoModel extraction → Mid-Product Excel → FinMod projections → Final Excel

This script orchestrates the entire workflow:
1. Extract financial data from 10-K HTML filing
2. Generate Mid-Product Excel (bridge between automodel and finmod)
3. Run financial projections in finmod
4. Generate Final.xlsx with assumptions and projections
"""

import os
import sys
import re
import json
from pathlib import Path
from typing import Optional, Dict, List
import subprocess

# Add project paths to sys.path
sys.path.insert(0, str(Path.cwd()))
sys.path.insert(0, str(Path.cwd() / "automodel"))
sys.path.insert(0, str(Path.cwd() / "final-project_finmod-main"))

import pandas as pd
from io import StringIO
import yaml
import math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Import automodel modules
from automodel.src.extract.is_tidy import tidy_is
from automodel.src.map.map_to_coa import map_labels
from automodel.src.llm.ollama_client import infer_scale, _llm_chat

# Convert extracted values to millions for consistent downstream modeling/display
SCALE_FACTORS = {
    "units": 1.0 / 1_000_000.0,
    "thousands": 1.0 / 1_000.0,
    "millions": 1.0,
    "billions": 1_000.0,  # billions → millions
}

YEAR_RE = re.compile(r"(?<!\d)(20\d{2})(?!\d)", re.I)
BASE_DIR = Path(__file__).resolve().parent
LAST_SUMMARY_PATH = BASE_DIR / "automodel/data/interim/last_summary.csv"
LAST_META_PATH = BASE_DIR / "automodel/data/interim/last_meta.json"
USE_LLM_EXTRACTION = os.environ.get("USE_LLM_EXTRACTION", "0").lower() in {"1", "true", "yes", "on"}


def _heuristic_score(df: pd.DataFrame) -> int:
    """Basic heuristic score for income statement likelihood."""
    first_col = df.iloc[:, 0].astype(str).str.lower()
    has_revenue = any('net sales' in x or 'revenue' in x for x in first_col)
    has_cogs = any('cost of sales' in x or 'cogs' in x for x in first_col)
    has_gross = any('gross' in x for x in first_col)
    has_operating = any('operating' in x or 'ebitda' in x for x in first_col)
    has_net_income = any('net income' in x for x in first_col)
    return sum([has_revenue, has_cogs, has_gross, has_operating, has_net_income])


def _summarize_table(df: pd.DataFrame, idx: int) -> str:
    """Lightweight text summary of a table for LLM ranking."""
    col_headers = [str(c) for c in df.columns]
    first_rows = []
    for _, row in df.head(8).iterrows():
        first_rows.append("|".join(str(x) for x in row.tolist()))
    years = YEAR_RE.findall(" ".join(col_headers) + " " + df.head(4).astype(str).to_string())
    return (
        f"Table {idx}: shape={df.shape}, headers={col_headers}, years_in_header={years}, "
        f"first_rows={first_rows}"
    )


def _rank_tables(dfs: List[pd.DataFrame], use_llm: bool) -> List[int]:
    """Return a ranked list of table indices (best first)."""
    scored = [( _heuristic_score(df), idx) for idx, df in enumerate(dfs)]
    scored_sorted = [idx for _, idx in sorted(scored, key=lambda t: t[0], reverse=True)]
    ranked = scored_sorted[:]

    if use_llm:
        top_candidates = scored_sorted[:6]
        summaries = [_summarize_table(dfs[idx], idx) for idx in top_candidates]
        prompt = (
            "You are selecting the INCOME STATEMENT table from scraped HTML tables. "
            "Pick the one that most likely represents a P&L (Revenue/Net sales, Cost of sales/COGS, "
            "Gross profit, Operating income, Net income). "
            "Respond with ONLY the table index number.\n\n"
            + "\n\n".join(summaries)
        )
        try:
            resp = _llm_chat(
                [
                    {"role": "system", "content": "Select the income statement table index."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.0,
            )
            m = re.search(r"(\d+)", resp)
            if m:
                idx = int(m.group(1))
                if 0 <= idx < len(dfs):
                    # put LLM choice first
                    ranked = [idx] + [i for i in scored_sorted if i != idx]
        except Exception as e:
            print(f"[WARN] LLM table selection failed ({e}); falling back to heuristic.")

    return ranked or [0]


def _llm_extract_is(raw_html: str) -> pd.DataFrame:
    """
    Ask LLM to extract income statement figures from raw HTML.
    Returns DataFrame with label_raw, year, value, scale_hint.
    """
    # Cap payload size to avoid OOM/timeouts in constrained environments
    RAW_LIMIT = 20000  # characters
    html_payload = raw_html[:RAW_LIMIT]

    prompt = (
        "Extract a consolidated income statement (P&L) from the given HTML content. "
        "Output ONLY JSON with fields: "
        "{ \"unit\": \"millions|thousands|billions|ones\", "
        "\"items\": [ {\"label\": \"Revenue\", \"year\": 2024, \"value\": 1234.56}, ... ] }. "
        "Values should be in the stated unit (e.g., if unit=millions, values are millions). "
        "Include at least Revenue, Cost of revenue/COGS, Gross profit, Operating income, Income before taxes, Net income for the last 3 reported years if available."
    )
    try:
        resp = _llm_chat(
            [
                {"role": "system", "content": "You extract structured financials from HTML."},
                {"role": "user", "content": prompt + "\nHTML:\n" + html_payload},
            ],
            temperature=0.0,
        )
        txt = resp.strip()
        if txt.startswith("```"):
            parts = txt.split("```")
            if len(parts) >= 3:
                txt = parts[1]
            if txt.startswith("json"):
                txt = txt[len("json"):].strip()
        data = json.loads(txt)
        unit = str(data.get("unit", "millions")).lower()
        if unit not in {"millions", "thousands", "billions", "ones"}:
            unit = "millions"
        items = data.get("items", [])
        rows = []
        for item in items:
            label = str(item.get("label", "")).strip()
            year = int(item.get("year", 0))
            val = float(item.get("value", 0))
            if label and 1900 < year < 2100:
                rows.append({"label_raw": label, "year": year, "value": val, "scale_hint": unit})
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["label_raw", "year", "value", "scale_hint"])
    except Exception as e:
        print(f"[WARN] LLM extraction failed: {e}")
        return pd.DataFrame(columns=["label_raw", "year", "value", "scale_hint"])


def _map_and_save(
    t: pd.DataFrame,
    hdr_blob: str,
    use_llm_tables: bool,
    table_idx: int,
    skip_llm: bool
) -> Optional[Path]:
    """
    Common mapping/validation/save path for extracted tidy data.
    Returns csv path or None if validation fails.
    """
    if t is None or t.empty:
        return None

    hint = t["scale_hint"].iloc[0] if "scale_hint" in t.columns and len(t) else "units"
    if len(t):
        try:
            sample_vals = (
                t["value"].dropna().astype(float).sample(min(6, len(t))).tolist()
            )
        except Exception:
            sample_vals = []
    else:
        sample_vals = []
    inferred = infer_scale(hdr_blob[:400], sample_vals) if hint == "units" else hint
    factor = SCALE_FACTORS.get(inferred, 1.0)
    t["value"] = t["value"] * factor

    all_tidy = t[["label_raw", "year", "value"]].copy()
    mapped = map_labels(all_tidy, BASE_DIR / "automodel/configs/mappings.yaml")

    with open(BASE_DIR / "automodel/configs/coa.yaml", "r") as f:
        coa_candidates = list((yaml.safe_load(f) or {}).keys())

    unm = mapped["coa"].isna()
    if unm.any():
        if skip_llm:
            print("Skipping LLM mapping. Unmapped labels will remain empty.")
        else:
            from automodel.src.llm.ollama_client import map_label_to_coa
            uniq = mapped.loc[unm, "label_raw"].dropna().unique().tolist()
            try:
                llm_map = {lbl: map_label_to_coa(lbl, coa_candidates) for lbl in uniq}
            except Exception as e:
                print(f"[WARN] LLM mapping failed ({e}); continuing without LLM.")
                llm_map = {}
            if llm_map:
                mapped.loc[unm, "coa"] = mapped.loc[unm, "label_raw"].map(llm_map)

    # Keep only P&L lines
    REVENUE_COAS = {"Revenue", "Interest Income"}
    EXPENSE_COAS = {
        "COGS",
        "Sales & Marketing",
        "Research & Development",
        "General & Administrative",
        "Depreciation & Amortization",
        "Share-Based Compensation",
        "Interest Expense",
        "Income Tax Expense",
        "Other Income (Expense)",
    }
    PNL_TOTAL_COAS = REVENUE_COAS.union(EXPENSE_COAS).union(
        {
            "Gross Profit",
            "Operating Income (EBIT)",
            "Income Before Taxes",
            "Net Income",
        }
    )
    mapped = mapped[mapped["coa"].isin(PNL_TOTAL_COAS)].copy()

    # Sign normalization
    def _norm(row):
        v = float(row["value"])
        c = row.get("coa")
        if c in REVENUE_COAS:
            return abs(v)
        if c in EXPENSE_COAS:
            return -abs(v)
        return v
    mapped["value"] = mapped.apply(_norm, axis=1)

    view = (
        mapped.dropna(subset=["coa"])
        .groupby(["coa", "year"])["value"]
        .sum()
        .reset_index()
        .sort_values(["coa", "year"])
    )

    # Validation: require revenue and cogs for at least 2 years and gross profit consistency
    rev = view[view["coa"] == "Revenue"].set_index("year")["value"] if not view.empty else {}
    cogs = view[view["coa"] == "COGS"].set_index("year")["value"] if not view.empty else {}
    gp = view[view["coa"] == "Gross Profit"].set_index("year")["value"] if not view.empty else {}
    if len(rev) < 2 or len(cogs) < 2:
        print(f"[WARN] Table {table_idx} rejected: insufficient revenue/COGS data.")
        return None
    bad_years = []
    for y in gp.index:
        r = rev.get(y, 0)
        c = cogs.get(y, 0)
        g = gp.get(y, 0)
        if abs((r + c) - g) > max(1e-3, 0.02 * max(abs(g), 1)):
            bad_years.append(y)
    if bad_years and len(bad_years) == len(gp.index):
        print(f"[WARN] Table {table_idx} rejected: gross profit mismatch.")
        return None

    outdir = BASE_DIR / "automodel/data/interim"
    outdir.mkdir(parents=True, exist_ok=True)
    all_tidy.to_csv(outdir / "IS_tidy_best.csv", index=False)
    csv_path = outdir / "IS_tidy_mapped_best_llm.csv"
    mapped.to_csv(csv_path, index=False)
    view.to_csv(LAST_SUMMARY_PATH, index=False)
    print("\nSummary by COA & year:")
    print(view.to_string(index=False))
    print(f"✅ Extracted financial data to {csv_path}")
    try:
        meta = {
            "table_index": table_idx,
            "use_llm_for_tables": use_llm_tables,
            "units": inferred,
            "years": sorted(view["year"].unique().tolist()),
        }
        LAST_META_PATH.write_text(json.dumps(meta, indent=2))
    except Exception as e:
        print(f"[WARN] Could not save metadata: {e}")
    return csv_path


def custom_tidy_is(df: pd.DataFrame) -> pd.DataFrame:
    """
    Custom tidying for income statement tables from pd.read_html.
    Handles tables with complex headers and numeric columns.
    """
    detected_scale = "units"
    # Inspect top rows for scale hints before dropping
    top_blob = " ".join(df.head(6).astype(str).stack().tolist()).lower()
    if "million" in top_blob:
        detected_scale = "millions"
    elif "thousand" in top_blob:
        detected_scale = "thousands"
    elif "billion" in top_blob:
        detected_scale = "billions"

    # Drop completely empty columns to simplify year detection
    df = df.dropna(axis=1, how="all").reset_index(drop=True)

    # Remove header rows that are just the "(in thousands...)" note
    header_note_re = re.compile(r"in thousands", re.I)
    rows_to_drop = []
    for i in range(min(4, len(df))):
        row_blob = " ".join(df.iloc[i].astype(str).tolist())
        if header_note_re.search(row_blob.lower()):
            # Keep the row if it also contains explicit years (it might be the year header)
            if not YEAR_RE.search(row_blob):
                rows_to_drop.append(i)
    if rows_to_drop:
        df = df.drop(index=rows_to_drop).reset_index(drop=True)

    # Find year columns
    year_cols = {}
    header_blob = " ".join([str(c) for c in df.columns]) + " " + df.head(6).astype(str).to_string()
    
    # Extract unique years from header
    years_found = []
    for mo in YEAR_RE.finditer(header_blob):
        y = int(mo.group(1))
        if y not in years_found:
            years_found.append(y)
    
    if not years_found:
        return pd.DataFrame(columns=["label_raw", "year", "value", "scale_hint"])
    
    # Try to map columns to years.
    # Strategy: look at the first few rows per column for a year token, then map that
    # year to the nearest numeric column (current or next few columns).
    year_map: Dict[int, int] = {}

    # 1) Try to find the row with the most year tokens and map each year token
    # to the nearest numeric-dominant column to its right.
    year_row_idx = None
    max_years_in_row = 0
    for i in range(min(8, len(df))):
        yrs = YEAR_RE.findall(" ".join(df.iloc[i].astype(str).tolist()))
        if len(yrs) > max_years_in_row:
            max_years_in_row = len(yrs)
            year_row_idx = i

    if year_row_idx is not None and max_years_in_row > 0:
        row_vals = df.iloc[year_row_idx]
        for idx, val in row_vals.items():
            m = YEAR_RE.search(str(val))
            if not m:
                continue
            year = int(m.group(1))
            if year in year_map.values():
                continue
            # search to the right for a numeric-dominant column
            numeric_target = None
            for j in range(df.columns.get_loc(idx) + 1, min(df.columns.get_loc(idx) + 4, len(df.columns))):
                col_label = df.columns[j]
                try:
                    # Clean numbers similarly to to_number for better detection
                    def _clean_num(x):
                        s = str(x).strip().replace(",", "").replace("$", "")
                        if s.startswith("(") and s.endswith(")"):
                            s = "-" + s[1:-1]
                        try:
                            return float(s)
                        except Exception:
                            return None
                    cleaned = df.loc[year_row_idx + 1 :, col_label].map(_clean_num)
                    valid = cleaned.dropna()
                    numeric_ratio = (len(valid) / max(1, len(cleaned)))
                    if numeric_ratio > 0.2:  # has numeric data
                        numeric_target = col_label
                        break
                except Exception:
                    continue
            if numeric_target is None:
                numeric_target = idx
            year_map[numeric_target] = year

    # 2) If still empty, fall back to per-column header scan
    if not year_map:
        for idx, col_label in enumerate(df.columns):
            header_bits = [str(col_label)]
            for r in range(min(6, len(df))):
                header_bits.append(str(df.iloc[r, idx]))
            blob = " ".join(header_bits)
            m = YEAR_RE.search(blob)
            if not m:
                continue
            year = int(m.group(1))
            if year in year_map.values():
                continue
            # Pick this column or the next couple columns that actually contain numbers
            numeric_target = None
            for j in range(idx, min(idx + 3, len(df.columns))):
                try:
                    col_vals = pd.to_numeric(df.iloc[:, j], errors='coerce').dropna()
                    if len(col_vals) > 0:
                        numeric_target = df.columns[j]
                        break
                except Exception:
                    continue
            numeric_target = numeric_target if numeric_target is not None else col_label
            year_map[numeric_target] = year

    if not year_map and years_found:
        # Fallback: use rightmost numeric-ish columns matched to detected years
        numeric_cols = []
        for col_label in df.columns:
            try:
                col_vals = pd.to_numeric(df[col_label], errors='coerce').dropna()
                if len(col_vals) > 0:
                    numeric_cols.append(col_label)
            except Exception:
                pass
        numeric_cols = numeric_cols[-len(years_found):]
        year_map = {col_label: int(y) for col_label, y in zip(numeric_cols, years_found)}

    if not year_map:
        return pd.DataFrame(columns=["label_raw", "year", "value", "scale_hint"])

    # Choose the label column: first non-numeric-heavy column (fallback to first)
    label_col = df.columns[0]
    best_label_col = None
    best_score = -1
    for col in df.columns:
        series = df[col].dropna().astype(str)
        if series.empty:
            continue
        texty = sum(1 for v in series.head(8) if not re.fullmatch(r"-?\\d+(\\.\\d+)?", str(v).replace(',', '')))
        if texty > best_score:
            best_score = texty
            best_label_col = col
    if best_label_col is not None:
        label_col = best_label_col

    # Extract data
    rows = []
    
    def to_number(x):
        s = str(x).strip().replace(",", "").replace("$", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        if s in {"", "—", "–", "-", "nan", "NaN", "N/A"}:
            return None
        try:
            return float(s)
        except:
            return None
    
    for _, r in df.iterrows():
        label = str(r.get(label_col, "")).strip()
        if not label or pd.isna(label) or label.lower() in {'nan', ''}:
            continue
        if header_note_re.search(label.lower()):
            continue
        
        for col_idx, year in year_map.items():
            if col_idx not in r.index:
                continue
            val = to_number(r[col_idx])
            if val is not None:
                # Use detected scale hint so infer_scale can refine only when unknown
                rows.append({"label_raw": label, "year": year, "value": val, "scale_hint": detected_scale})
    
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["label_raw", "year", "value", "scale_hint"])


def step1_extract_from_html(html_path: Path, skip_llm: bool = True) -> Path:
    """
    Step 1: Run automodel extraction on 10-K HTML filing.
    
    Args:
        html_path: Path to 10-K HTML file
        skip_llm: If True, skip LLM-based label mapping
        
    Returns:
        Path to the mapped CSV output (IS_tidy_mapped_best_llm.csv)
    """
    print("\n" + "="*60)
    print("STEP 1: Extracting financial data from 10-K HTML")
    print("="*60)
    
    if not html_path.exists():
        raise FileNotFoundError(f"HTML file not found: {html_path}")
    
    raw = html_path.read_text(errors="ignore")
    
    # First try LLM-based extraction if enabled
    if not skip_llm and USE_LLM_EXTRACTION:
        print("[INFO] Trying LLM-based extraction first...")
        t_llm = _llm_extract_is(raw)
        if t_llm is not None and not t_llm.empty:
            csv_path = _map_and_save(t_llm, raw[:400], use_llm_tables=True, table_idx=-1, skip_llm=skip_llm)
            if csv_path:
                return csv_path
        print("[WARN] LLM extraction failed or returned empty; falling back to table parsing.")
    
    # Parse all tables from HTML
    try:
        dfs = pd.read_html(StringIO(raw))
    except Exception as e:
        raise SystemExit(f"pd.read_html failed: {type(e).__name__}: {e}")
    
    if not dfs:
        raise SystemExit("No tables found in HTML.")
    
    print(f"Found {len(dfs)} tables in HTML")
    
    # Detect income statement table using heuristics + optional LLM (if not skipped)
    use_llm_tables = not skip_llm
    candidate_indices = _rank_tables(dfs, use_llm=use_llm_tables)
    good = False
    for cand in candidate_indices:
        is_df_raw = dfs[cand]
        is_idx = cand
        print(f"Trying table index {cand} as the income statement candidate.")
        t = custom_tidy_is(is_df_raw)
        if t is None or t.empty:
            print("[WARN] Custom tidy returned empty; trying standard tidy_is...")
            t = tidy_is(is_df_raw)
        if t is None or t.empty:
            continue
        csv_path = _map_and_save(t, " ".join(map(str, is_df_raw.columns)).lower(), use_llm_tables, is_idx, skip_llm)
        if csv_path:
            good = True
            return csv_path

    raise SystemExit("No valid income statement table found after validation.")


def step2_create_mid_product(
    csv_path: Path,
    template_path: Path,
    output_path: Path,
    company_name: str = "Company"
) -> Path:
    """
    Step 2: Convert automodel CSV to finmod-compatible Mid-Product Excel.
    
    Args:
        csv_path: Path to IS_tidy_mapped_best_llm.csv
        template_path: Path to Baseline IS.xlsx template
        output_path: Path to write Mid-Product.xlsx
        company_name: Company name for display
        
    Returns:
        Path to created Mid-Product Excel file
    """
    print("\n" + "="*60)
    print("STEP 2: Creating Mid-Product Excel")
    print("="*60)
    
    # Read the mapped data
    df = pd.read_csv(csv_path)
    
    # Remove rows without COA mapping
    df = df[df['coa'].notna()].copy()
    
    # Aggregate by COA and year
    df_pivot = df.groupby(['coa', 'year'])['value'].sum().reset_index()
    
    # Load the template workbook (with data_only=False to see formulas)
    wb = load_workbook(template_path, data_only=False)
    ws = wb.active
    
    # Update company name
    ws['D1'] = company_name
    
    # Get the year column mapping - handle both direct values and formulas.
    # Instead of relying on the template's starting year, anchor on the earliest
    # year present in the data so we don't drop older periods (e.g., when data
    # includes 2022 but the template header shows 2023).
    year_map = {}
    base_year = None

    # Anchor the template to the earliest data year and fill sequentially across projections
    data_years = sorted(df_pivot['year'].unique().tolist()) if not df_pivot.empty else []
    if data_years:
        base_year = data_years[0]
        # Fill all columns E-M with sequential years so projections align after actuals
        for offset in range(0, 9):  # E to M inclusive
            year = base_year + offset
            col_idx = 5 + offset
            if col_idx < 14:
                year_map[year] = col_idx
                ws.cell(row=4, column=col_idx).value = year
    else:
        # Fallback to whatever is present in the template
        for col_idx in range(5, 14):  # E to M (cols 5-13)
            cell = ws.cell(row=4, column=col_idx)
            if cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                try:
                    year = int(cell.value)
                    if 2000 <= year <= 2100:
                        year_map[year] = col_idx
                except (ValueError, TypeError):
                    pass
    
    if not year_map:
        raise ValueError("Could not find year columns in template")
    
    # Mapping from extracted COA to template rows
    coa_mapping = {
        'Revenue': 'Revenue',
        'COGS': 'COGS',
        'Gross Profit': 'Gross Profit',
        'Sales & Marketing': 'SG&A',
        'General & Administrative': 'SG&A',
        'Research & Development': 'R&D',
        'Depreciation & Amortization': 'R&D',
        'Share-Based Compensation': 'R&D',
        'Operating Income (EBIT)': 'Organic EBITDA',
        'Interest Expense': 'Other Income',
        'Interest Income': 'Other Income',
        'Other Income (Expense)': 'Other Income',
        'Income Tax Expense': 'Other Income',
        'Income Before Taxes': 'Organic EBITDA',
        'Net Income': 'Total EBITDA',
    }

    # Units note from metadata if available
    units = "millions"
    if LAST_META_PATH.exists():
        try:
            meta = json.loads(LAST_META_PATH.read_text())
            units = meta.get("units", units)
        except Exception:
            pass
    # Write units note in a safe, unused cell to preserve P&L header (e.g., B2)
    ws.cell(row=2, column=2, value=f"Units: {units}")

    # Find template row indices
    template_rows = {}
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)  # Column C
        if cell.value:
            label = str(cell.value).strip()
            for target in set(coa_mapping.values()):
                if label.lower() == target.lower():
                    template_rows[target] = row_idx
                    break
    
    # Aggregate extracted data
    aggregated = {}
    for _, row in df_pivot.iterrows():
        coa = row['coa']
        year = int(row['year'])
        value = row['value']
        
        template_label = coa_mapping.get(coa)
        if template_label:
            if template_label not in aggregated:
                aggregated[template_label] = {}
            
            if year not in aggregated[template_label]:
                aggregated[template_label][year] = 0
            aggregated[template_label][year] += value
    
    # Fill template with data
    for template_label, year_values in aggregated.items():
        if template_label not in template_rows:
            continue
        
        row_idx = template_rows[template_label]
        for year, value in year_values.items():
            if year not in year_map:
                continue
            col_idx = year_map[year]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.number_format = "#,##0;(#,##0)"
            cell.font = Font(color="1F4E79")  # blue tone for actuals

    # Inject formulas for Gross Profit, Organic EBITDA, Total EBITDA across all columns
    gp_row = template_rows.get("Gross Profit")
    rev_row = template_rows.get("Revenue")
    cogs_row = template_rows.get("COGS")
    sgna_row = template_rows.get("SG&A")
    rnd_row = template_rows.get("R&D")
    organic_row = template_rows.get("Organic EBITDA")
    other_row = template_rows.get("Other Income")
    total_row = template_rows.get("Total EBITDA")
    capex_row = template_rows.get("Capex")

    def _coord(r, c):
        return f"{get_column_letter(c)}{r}"

    # Assumption cells (fixed)
    ASSUMP = {
        "rev": "Q5",
        "cogs": "Q6",
        "sgna": "Q7",
        "rnd": "Q8",
        "capex": "Q9",
        "other": "Q10",
    }

    # Format all cells to Arial 12
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.font = Font(name="Arial", size=12, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)

    actual_years = data_years if data_years else []
    max_actual_year = max(actual_years) if actual_years else None

    for year, col_idx in year_map.items():
        # Projected revenue/COGS/SG&A/R&D/Other/Capex: prior year * (1 + assumption) only for projected years
        if max_actual_year and year > max_actual_year:
            prior_year = year - 1
            prior_col = year_map.get(prior_year)
            if prior_col:
                if rev_row:
                    ws.cell(row=rev_row, column=col_idx).value = f"={_coord(rev_row, prior_col)}*(1+{ASSUMP['rev']})"
                if cogs_row:
                    ws.cell(row=cogs_row, column=col_idx).value = f"={_coord(cogs_row, prior_col)}*(1+{ASSUMP['cogs']})"
                if sgna_row:
                    ws.cell(row=sgna_row, column=col_idx).value = f"={_coord(sgna_row, prior_col)}*(1+{ASSUMP['sgna']})"
                if rnd_row:
                    ws.cell(row=rnd_row, column=col_idx).value = f"={_coord(rnd_row, prior_col)}*(1+{ASSUMP['rnd']})"
                if other_row:
                    ws.cell(row=other_row, column=col_idx).value = f"={_coord(other_row, prior_col)}*(1+{ASSUMP['other']})"
                if capex_row:
                    ws.cell(row=capex_row, column=col_idx).value = f"={_coord(capex_row, prior_col)}*(1+{ASSUMP['capex']})"

        # Gross Profit, Organic EBITDA, Total EBITDA formulas
        if rev_row and cogs_row and gp_row:
            ws.cell(row=gp_row, column=col_idx).value = f"={_coord(rev_row,col_idx)}+{_coord(cogs_row,col_idx)}"
            ws.cell(row=gp_row, column=col_idx).number_format = "#,##0;(#,##0)"
        if gp_row and sgna_row and rnd_row and organic_row:
            ws.cell(row=organic_row, column=col_idx).value = (
                f"={_coord(gp_row,col_idx)}-{_coord(sgna_row,col_idx)}-{_coord(rnd_row,col_idx)}"
            )
            ws.cell(row=organic_row, column=col_idx).number_format = "#,##0;(#,##0)"
        if organic_row and other_row and total_row:
            ws.cell(row=total_row, column=col_idx).value = f"={_coord(organic_row,col_idx)}+{_coord(other_row,col_idx)}"
            ws.cell(row=total_row, column=col_idx).number_format = "#,##0;(#,##0)"

    # % rows link to assumptions for projected years
    pct_rows = {
        "cogs": template_rows.get("COGS"),
        "sgna": template_rows.get("SG&A"),
        "rnd": template_rows.get("R&D"),
        "other": template_rows.get("Other Income"),
    }
    for key, row_idx in pct_rows.items():
        if not row_idx:
            continue
        pct_row = row_idx + 1  # the % row is immediately below the value row
        for year, col_idx in year_map.items():
            # Only set for projected years (beyond actual years)
            if max_actual_year and year > max_actual_year:
                ws.cell(row=pct_row, column=col_idx).value = f"={ASSUMP[key]}"
    
    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Created Mid-Product: {output_path}")
    return output_path


def step3_run_finmod_projections(
    mid_product_path: Path,
    output_path: Path
) -> Path:
    """
    Step 3: Run finmod projections on Mid-Product to generate Final.xlsx.
    
    Args:
        mid_product_path: Path to Mid-Product.xlsx
        output_path: Path to write Final.xlsx
        
    Returns:
        Path to created Final.xlsx file
    """
    print("\n" + "="*60)
    print("STEP 3: Running FinMod Projections")
    print("="*60)
    
    if not mid_product_path.exists():
        raise FileNotFoundError(f"Mid-Product not found: {mid_product_path}")
    
    # Run finmod
    cmd = [
        sys.executable, "-m", "final-project_finmod-main.src.finmod.main",
        "--file", str(mid_product_path),
        "--output-xlsx", str(output_path)
    ]
    
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True
    )
    
    if result.returncode != 0:
        print("STDERR:", result.stderr)
        raise RuntimeError(f"FinMod projection failed: {result.stderr}")
    
    print("STDOUT:", result.stdout)
    
    if not output_path.exists():
        raise RuntimeError("FinMod did not produce expected output Excel")
    
    # Post-process final workbook to apply formulas/formatting
    try:
        _apply_projection_formulas(output_path)
    except Exception as e:
        print(f"[WARN] Could not apply projection formulas: {e}")

    print(f"✅ Generated Final Excel: {output_path}")
    # Add validation sheet if summary is available
    try:
        if LAST_SUMMARY_PATH.exists():
            _attach_validation_sheet(output_path, LAST_SUMMARY_PATH)
    except Exception as e:
        print(f"[WARN] Could not attach validation sheet: {e}")
    return output_path


def _attach_validation_sheet(final_path: Path, summary_path: Path) -> None:
    """
    Add a 'Validation' sheet to the final Excel showing extracted P&L and simple checks.
    """
    if not final_path.exists() or not summary_path.exists():
        return

    df = pd.read_csv(summary_path)
    if df.empty or "coa" not in df.columns:
        return

    pivot = (
        df.pivot_table(index="coa", columns="year", values="value", aggfunc="sum")
        .reindex(["Revenue", "COGS", "Gross Profit", "Operating Income (EBIT)", "Income Before Taxes", "Net Income",
                  "General & Administrative", "Sales & Marketing", "Research & Development"])
    )

    wb = load_workbook(final_path)
    if "Validation" in wb.sheetnames:
        ws = wb["Validation"]
        wb.remove(ws)
    ws = wb.create_sheet("Validation")

    ws["A1"] = "Extracted P&L"
    # Metadata
    units = "millions"
    if LAST_META_PATH.exists():
        try:
            meta = json.loads(LAST_META_PATH.read_text())
            ws["A2"] = f"Table index: {meta.get('table_index')}, LLM table select: {meta.get('use_llm_for_tables')}"
            units = meta.get("units", units)
            ws["A3"] = f"Units: {units}"
        except Exception:
            pass
    # headers
    years = sorted([int(y) for y in pivot.columns if pd.notna(y)]) if pivot.columns.size else []
    for j, year in enumerate(years, start=2):
        ws.cell(row=2, column=j, value=year)

    # rows
    label_map = {
        "Revenue": "Revenue",
        "COGS": "COGS",
        "Gross Profit": "Gross Profit (calc)",
        "Operating Income (EBIT)": "Operating Income",
        "Income Before Taxes": "Income Before Taxes",
        "Net Income": "Net Income",
    }

    def fmt_number(cell):
        cell.number_format = "#,##0;(#,##0)"
        return cell

    def get_val(label: str, year: int) -> float:
        if label in pivot.index and year in pivot.columns:
            try:
                return float(pivot.loc[label, year])
            except Exception:
                return 0.0
        return 0.0

    row_idx = 3
    for coa_key, display in label_map.items():
        ws.cell(row=row_idx, column=1, value=display)
        for j, year in enumerate(years, start=2):
            val = pivot.at[coa_key, year] if (coa_key in pivot.index and year in pivot.columns) else None
            if pd.notna(val):
                fmt_number(ws.cell(row=row_idx, column=j, value=val))
        row_idx += 1

    # Simple check: Gross Profit vs Revenue + COGS
    ws.cell(row=row_idx, column=1, value="Check: Gross = Revenue + COGS (difference)")
    for j, year in enumerate(years, start=2):
        rev = get_val("Revenue", year)
        cogs = get_val("COGS", year)
        gp = get_val("Gross Profit", year)
        diff = (rev + cogs) - gp
        c = fmt_number(ws.cell(row=row_idx, column=j, value=diff))
        if abs(diff) > 1e-2:
            c.font = c.font.copy(color="FF0000")
    row_idx += 2

    # Operating income check: Revenue + COGS + SG&A + R&D vs Operating Income
    ws.cell(row=row_idx, column=1, value="Check: Operating Income vs derived (difference)")
    for j, year in enumerate(years, start=2):
        rev = get_val("Revenue", year)
        cogs = get_val("COGS", year)
        gna = get_val("General & Administrative", year)
        sgna = get_val("Sales & Marketing", year)
        rnd = get_val("Research & Development", year)
        derived = rev + cogs + gna + sgna + rnd
        oi = get_val("Operating Income (EBIT)", year)
        diff = derived - oi
        c = fmt_number(ws.cell(row=row_idx, column=j, value=diff))
        if abs(diff) > 1e-2:
            c.font = c.font.copy(color="FF0000")
    row_idx += 2

    # Net income check: Income Before Taxes + Income Tax Expense vs Net Income
    ws.cell(row=row_idx, column=1, value="Check: Net Income vs PBT + Tax (difference)")
    for j, year in enumerate(years, start=2):
        pbt = get_val("Income Before Taxes", year)
        tax = get_val("Income Tax Expense", year)
        ni = get_val("Net Income", year)
        diff = (pbt + tax) - ni
        c = fmt_number(ws.cell(row=row_idx, column=j, value=diff))
        if abs(diff) > 1e-2:
            c.font = c.font.copy(color="FF0000")
    wb.save(final_path)


def _apply_projection_formulas(final_path: Path) -> None:
    """
    Rewrite projected-year cells to formulas tied to assumptions,
    and recalc GP/EBITDA rows. Also normalize font to Arial 12.
    """
    if not final_path.exists():
        return
    wb = load_workbook(final_path, data_only=False)
    ws = wb.active

    # Build year map from header row (assume row 4, columns E-M)
    year_map = {}
    base_row = 4
    for col_idx in range(5, 14):
        cell = ws.cell(row=base_row, column=col_idx)
        try:
            year = int(cell.value)
            year_map[year] = col_idx
        except Exception:
            continue
    if not year_map:
        wb.save(final_path)
        return

    # Locate rows by label in column C
    labels = {}
    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=3).value
        if isinstance(val, str):
            key = val.strip().lower()
            labels[key] = row_idx

    def get_row(label):
        return labels.get(label.lower())

    rev_row = get_row("Revenue")
    cogs_row = get_row("COGS")
    gp_row = get_row("Gross Profit")
    sgna_row = get_row("SG&A")
    rnd_row = get_row("R&D")
    organic_row = get_row("Organic EBITDA")
    other_row = get_row("Other Income")
    total_row = get_row("Total EBITDA")
    capex_row = get_row("Capex")

    # Assumption cells
    ASSUMP = {
        "rev": "Q5",
        "cogs": "Q6",
        "sgna": "Q7",
        "rnd": "Q8",
        "capex": "Q9",
        "other": "Q10",
    }

    # Normalize font
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.font = Font(name="Arial", size=12, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)

    def _coord(r, c):
        return f"{get_column_letter(c)}{r}"

    actual_years = sorted([y for y in year_map.keys()])[:3]  # assume first 3 as actuals
    max_actual_year = max(actual_years) if actual_years else None

    for year, col_idx in year_map.items():
        if max_actual_year and year > max_actual_year:
            prior_year = year - 1
            prior_col = year_map.get(prior_year)
            if prior_col:
                if rev_row:
                    ws.cell(row=rev_row, column=col_idx).value = f"={_coord(rev_row, prior_col)}*(1+{ASSUMP['rev']})"
                if cogs_row:
                    ws.cell(row=cogs_row, column=col_idx).value = f"={_coord(cogs_row, prior_col)}*(1+{ASSUMP['cogs']})"
                if sgna_row:
                    ws.cell(row=sgna_row, column=col_idx).value = f"={_coord(sgna_row, prior_col)}*(1+{ASSUMP['sgna']})"
                if rnd_row:
                    ws.cell(row=rnd_row, column=col_idx).value = f"={_coord(rnd_row, prior_col)}*(1+{ASSUMP['rnd']})"
                if other_row:
                    ws.cell(row=other_row, column=col_idx).value = f"={_coord(other_row, prior_col)}*(1+{ASSUMP['other']})"
                if capex_row:
                    ws.cell(row=capex_row, column=col_idx).value = f"={_coord(capex_row, prior_col)}*(1+{ASSUMP['capex']})"

        # GP / EBITDA formulas all years
        if rev_row and cogs_row and gp_row:
            ws.cell(row=gp_row, column=col_idx).value = f"={_coord(rev_row,col_idx)}+{_coord(cogs_row,col_idx)}"
            ws.cell(row=gp_row, column=col_idx).number_format = "#,##0;(#,##0)"
        if gp_row and sgna_row and rnd_row and organic_row:
            ws.cell(row=organic_row, column=col_idx).value = (
                f"={_coord(gp_row,col_idx)}-{_coord(sgna_row,col_idx)}-{_coord(rnd_row,col_idx)}"
            )
            ws.cell(row=organic_row, column=col_idx).number_format = "#,##0;(#,##0)"
        if organic_row and other_row and total_row:
            ws.cell(row=total_row, column=col_idx).value = f"={_coord(organic_row,col_idx)}+{_coord(other_row,col_idx)}"
            ws.cell(row=total_row, column=col_idx).number_format = "#,##0;(#,##0)"

    # % rows for projected years
    pct_rows = {
        "cogs": cogs_row,
        "sgna": sgna_row,
        "rnd": rnd_row,
        "other": other_row,
    }
    for key, row_idx in pct_rows.items():
        if not row_idx:
            continue
        pct_row = row_idx + 1
        for year, col_idx in year_map.items():
            if max_actual_year and year > max_actual_year:
                ws.cell(row=pct_row, column=col_idx).value = f"={ASSUMP[key]}"

    wb.save(final_path)


def main(
    html_path: Path,
    mid_product_path: Optional[Path] = None,
    final_output_path: Optional[Path] = None,
    company_name: str = "Company",
    skip_llm: bool = True
):
    """
    Run the complete pipeline from 10-K HTML to Final Excel projections.
    
    Args:
        html_path: Path to 10-K HTML filing
        mid_product_path: Path for Mid-Product.xlsx (default: Mid-Product.xlsx)
        final_output_path: Path for Final.xlsx (default: Final.xlsx)
        company_name: Company name for display
        skip_llm: If True, skip LLM-based label mapping
    """
    
    print("\n" + "#"*60)
    print("# UNIFIED 10-K FINANCIAL ANALYSIS PIPELINE")
    print("#"*60)
    
    html_path = Path(html_path)
    mid_product_path = Path(mid_product_path or "Mid-Product.xlsx")
    final_output_path = Path(final_output_path or "Final.xlsx")
    template_path = Path("final-project_finmod-main/Inputs_Historical/Baseline IS.xlsx")
    
    try:
        # Step 1: Extract
        csv_path = step1_extract_from_html(html_path, skip_llm=skip_llm)
        
        # Step 2: Create Mid-Product
        step2_create_mid_product(
            csv_path,
            template_path,
            mid_product_path,
            company_name=company_name
        )
        
        # Step 3: Run projections
        step3_run_finmod_projections(mid_product_path, final_output_path)
        
        print("\n" + "#"*60)
        print("# PIPELINE COMPLETE!")
        print("#"*60)
        print(f"Mid-Product: {mid_product_path}")
        print(f"Final Output: {final_output_path}")
        print("#"*60 + "\n")
        
    except Exception as e:
        print(f"\n❌ Pipeline failed: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="End-to-end pipeline: 10-K HTML → Mid-Product.xlsx → Final.xlsx"
    )
    parser.add_argument(
        "--html",
        required=True,
        help="Path to 10-K HTML filing"
    )
    parser.add_argument(
        "--mid-product",
        default="Mid-Product.xlsx",
        help="Output path for Mid-Product.xlsx (default: Mid-Product.xlsx)"
    )
    parser.add_argument(
        "--final",
        default="Final.xlsx",
        help="Output path for Final.xlsx (default: Final.xlsx)"
    )
    parser.add_argument(
        "--company",
        default="Company",
        help="Company name for display (default: Company)"
    )
    parser.add_argument(
        "--use-llm",
        action="store_true",
        help="Use LLM for label mapping (default: disabled for speed)"
    )
    
    args = parser.parse_args()
    
    main(
        html_path=args.html,
        mid_product_path=args.mid_product,
        final_output_path=args.final,
        company_name=args.company,
        skip_llm=not args.use_llm
    )
