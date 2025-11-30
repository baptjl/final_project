"""
Utilities to infer assumptions from a simple income statement template and project
future periods using deterministic heuristics (no networked AI required).
"""

from __future__ import annotations

import math
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, MutableMapping, Tuple

from openpyxl import Workbook, load_workbook

Number = float


def _col_letters(cell_ref: str) -> str:
    """Return the column letters from a cell reference like 'C5'."""
    return "".join(ch for ch in cell_ref if ch.isalpha())


def _norm_label(text: str) -> str:
    """Normalize label text for loose matching (collapse whitespace, lower-case)."""
    return " ".join(text.split()).lower()


def load_grid_from_xlsx(path: Path) -> Dict[int, Dict[str, str]]:
    """
    Lightweight .xlsx reader that returns a sparse grid:
    {row_index: {column_letters: value}} using openpyxl to handle inline strings.
    """
    grid: Dict[int, Dict[str, str]] = defaultdict(dict)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_letters = cell.column_letter
            grid[cell.row][col_letters] = cell.value
    return grid


def _coerce_float(val: str | None) -> Number:
    try:
        return float(val) if val not in ("", None) else math.nan
    except ValueError:
        return math.nan


def _find_year_row(grid: Mapping[int, Mapping[str, str]]) -> Tuple[int, Dict[str, int]]:
    """Locate the header row containing year labels and return its mapping."""
    for row_idx, cols in sorted(grid.items()):
        year_map: Dict[str, int] = {}
        for col, val in cols.items():
            try:
                year_map[col] = int(val)
            except (TypeError, ValueError):
                continue
        if year_map:
            return row_idx, year_map
    raise ValueError("Could not locate a row with year labels in the workbook.")


def _find_row_by_label(
    grid: Mapping[int, Mapping[str, str]], label: str, label_col: str = "C"
) -> Mapping[str, str]:
    target = _norm_label(label)
    for _, cols in sorted(grid.items()):
        if _norm_label(str(cols.get(label_col, ""))) == target:
            return cols
    raise KeyError(f"Label '{label}' not found in column {label_col}.")


def _find_row_index_by_label(
    grid: Mapping[int, Mapping[str, str]], label: str, label_col: str = "C"
) -> int:
    target = _norm_label(label)
    for row_idx, cols in sorted(grid.items()):
        if _norm_label(str(cols.get(label_col, ""))) == target:
            return row_idx
    raise KeyError(f"Label '{label}' not found in column {label_col}.")


@dataclass
class IncomeStatement:
    years: List[int]
    revenue: Dict[int, Number]
    cogs: Dict[int, Number]
    sgna: Dict[int, Number]
    rnd: Dict[int, Number]
    other_income: Dict[int, Number]
    capex: Dict[int, Number]


def load_income_statement(path: Path) -> IncomeStatement:
    """Parse the baseline workbook into structured series by year."""
    grid = load_grid_from_xlsx(path)
    _, year_map = _find_year_row(grid)

    def extract(label: str) -> Dict[int, Number]:
        row = _find_row_by_label(grid, label)
        series: Dict[int, Number] = {}
        for col, year in year_map.items():
            val = _coerce_float(row.get(col))
            if not math.isnan(val):
                series[year] = val
        return series

    years_sorted = sorted(year_map.values())
    return IncomeStatement(
        years=years_sorted,
        revenue=extract("Revenue"),
        cogs=extract("COGS"),
        sgna=extract("SG&A"),
        rnd=extract("R&D"),
        other_income=extract("Other Income"),
        capex=extract("Capex"),
    )


def extract_all_series(path: Path) -> Tuple[List[int], str, Dict[str, Dict[int, Number]]]:
    """
    Extract all line-item series keyed by their label in column C.
    Returns (years list, revenue_label, series_map).
    """
    grid = load_grid_from_xlsx(path)
    _, year_map = _find_year_row(grid)
    years_sorted = sorted(year_map.values())

    series_map: Dict[str, Dict[int, Number]] = {}
    for _, cols in sorted(grid.items()):
        label = cols.get("C")
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        values: Dict[int, Number] = {}
        for col, year in year_map.items():
            val = _coerce_float(cols.get(col))
            if not math.isnan(val):
                values[year] = val
        if values:
            series_map[label] = values

    revenue_label = next((lbl for lbl in series_map if _norm_label(lbl) == _norm_label("Revenue")), None)
    if not revenue_label:
        raise ValueError("Could not locate a Revenue line in column C.")

    return years_sorted, revenue_label, series_map


def infer_dynamic_assumptions(
    years: List[int], revenue_label: str, series_map: Mapping[str, Mapping[int, Number]]
) -> Tuple[Number, Dict[str, Tuple[str, Number]]]:
    """
    Infer revenue CAGR and per-line projection strategy.
    Strategies per line:
      - ('ratio', pct_of_revenue)
      - ('cagr', growth_rate)
      - ('flat', last_value)
    """
    revenue = series_map[revenue_label]
    actual_years = sorted(revenue)
    if len(actual_years) < 2:
        raise ValueError("Need at least two revenue periods to infer growth.")
    first, last = actual_years[0], actual_years[-1]
    periods = len(actual_years) - 1
    revenue_cagr = (revenue[last] / revenue[first]) ** (1 / periods) - 1

    strategies: Dict[str, Tuple[str, Number]] = {}
    for label, series in series_map.items():
        if label == revenue_label:
            continue
        ratios = []
        vals = []
        for y in actual_years:
            rv = revenue.get(y)
            sv = series.get(y)
            if rv not in (None, 0, math.nan) and sv not in (None, math.nan):
                ratios.append(sv / rv)
                vals.append(sv)
        if len(ratios) >= 2:
            strategies[label] = ("ratio", sum(ratios) / len(ratios))
            continue
        if len(vals) >= 2 and vals[0] not in (None, 0):
            growth = (vals[-1] / vals[0]) ** (1 / (len(vals) - 1)) - 1
            strategies[label] = ("cagr", growth)
            continue
        if vals:
            strategies[label] = ("flat", vals[-1])
        else:
            strategies[label] = ("flat", 0.0)
    return revenue_cagr, strategies


def project_dynamic(
    years: List[int],
    revenue_label: str,
    series_map: Mapping[str, Mapping[int, Number]],
    revenue_cagr: Number,
    strategies: Mapping[str, Tuple[str, Number]],
) -> Dict[str, Dict[int, Number]]:
    """
    Project revenue and all other lines based on inferred strategies.
    """
    revenue = dict(series_map[revenue_label])
    actual_years = sorted(revenue)
    forecast_years = [y for y in years if y > actual_years[-1]]
    for year in forecast_years:
        prior = revenue[year - 1]
        revenue[year] = prior * (1 + revenue_cagr)

    projected: Dict[str, Dict[int, Number]] = {revenue_label: revenue}

    for label, series in series_map.items():
        if label == revenue_label:
            continue
        mode, param = strategies.get(label, ("flat", 0.0))
        full = dict(series)
        if mode == "ratio":
            for year in forecast_years:
                full[year] = revenue[year] * param
        elif mode == "cagr":
            # grow from last available value
            if series:
                last_year = max(series)
                last_val = series[last_year]
                for idx, year in enumerate(forecast_years, start=1):
                    full[year] = last_val * ((1 + param) ** idx)
        else:  # flat
            last_val = next(reversed(sorted(series.values()))) if series else 0.0
            for year in forecast_years:
                full[year] = last_val
        projected[label] = full

    return projected

@dataclass
class Assumptions:
    revenue_growth_cagr: Number
    cogs_pct: Number
    sgna_pct: Number
    rnd_pct: Number
    other_income_pct: Number
    capex_pct: Number


def infer_assumptions(is_model: IncomeStatement) -> Assumptions:
    """Derive growth and margin assumptions from historical actuals."""
    actual_years = sorted(is_model.revenue)
    if len(actual_years) < 2:
        raise ValueError("Need at least two actual periods to infer growth.")

    first, last = actual_years[0], actual_years[-1]
    periods = len(actual_years) - 1
    growth = (is_model.revenue[last] / is_model.revenue[first]) ** (1 / periods) - 1

    def avg_ratio(series: Mapping[int, Number]) -> Number:
        numerators = []
        for y in actual_years:
            rev = is_model.revenue.get(y, math.nan)
            val = series.get(y, math.nan)
            if not math.isnan(rev) and rev != 0 and not math.isnan(val):
                numerators.append(val / rev)
        return sum(numerators) / len(numerators) if numerators else 0.0

    return Assumptions(
        revenue_growth_cagr=growth,
        cogs_pct=avg_ratio(is_model.cogs),
        sgna_pct=avg_ratio(is_model.sgna),
        rnd_pct=avg_ratio(is_model.rnd),
        other_income_pct=avg_ratio(is_model.other_income),
        capex_pct=avg_ratio(is_model.capex),
    )


def project_statement(
    is_model: IncomeStatement, assumptions: Assumptions
) -> Dict[str, Dict[int, Number]]:
    """
    Return complete time series dicts with projections filled in for future years.

    EBITDA follows the pattern of the baseline: Revenue - COGS - SG&A + Other Income.
    """
    all_years = sorted(is_model.years)
    actual_years = sorted(is_model.revenue)
    forecast_years = [y for y in all_years if y > actual_years[-1]]

    revenue = dict(is_model.revenue)
    for year in forecast_years:
        prior = revenue[year - 1]
        revenue[year] = prior * (1 + assumptions.revenue_growth_cagr)

    def fill_pct(series: Mapping[int, Number], pct: Number) -> Dict[int, Number]:
        full = dict(series)
        for year in forecast_years:
            full[year] = revenue[year] * pct
        return full

    cogs = fill_pct(is_model.cogs, assumptions.cogs_pct)
    sgna = fill_pct(is_model.sgna, assumptions.sgna_pct)
    rnd = fill_pct(is_model.rnd, assumptions.rnd_pct)
    other_income = fill_pct(is_model.other_income, assumptions.other_income_pct)
    capex = fill_pct(is_model.capex, assumptions.capex_pct)

    gross_profit = {y: revenue[y] - cogs[y] for y in all_years}
    organic_ebitda = {y: revenue[y] - cogs[y] - sgna[y] + other_income[y] for y in all_years}
    total_ebitda = dict(organic_ebitda)

    return {
        "Revenue": revenue,
        "COGS": cogs,
        "Gross Profit": gross_profit,
        "SG&A": sgna,
        "R&D": rnd,
        "Other Income": other_income,
        "Capex": capex,
        "Organic EBITDA": organic_ebitda,
        "Total EBITDA": total_ebitda,
    }


def format_table(series_map: Mapping[str, Mapping[int, Number]], years: Iterable[int]) -> str:
    """Render a simple text table showing each line item by year."""
    years_list = list(sorted(years))
    header = ["Line Item"] + [str(y) for y in years_list]
    lines = ["\t".join(header)]
    for label, series in series_map.items():
        row = [label]
        for y in years_list:
            val = series.get(y)
            row.append(f"{val:,.2f}" if val is not None else "")
        lines.append("\t".join(row))
    return "\n".join(lines)


def write_projections_to_xlsx(
    output_path: Path,
    assumptions: Assumptions,
    projections: Mapping[str, Mapping[int, Number]],
    years: Iterable[int],
) -> None:
    """
    Write assumptions and projected line items to a new XLSX file.

    Sheet 1: Assumptions
    Sheet 2: Projected Income Statement
    """
    years_list = list(sorted(years))
    wb = Workbook()

    # Assumptions sheet
    sheet_assump = wb.active
    sheet_assump.title = "Assumptions"
    sheet_assump.append(["Assumption", "Value"])
    sheet_assump.append(["Revenue CAGR", f"{assumptions.revenue_growth_cagr:.4f}"])
    sheet_assump.append(["COGS % of revenue", f"{assumptions.cogs_pct:.4f}"])
    sheet_assump.append(["SG&A % of revenue", f"{assumptions.sgna_pct:.4f}"])
    sheet_assump.append(["R&D % of revenue", f"{assumptions.rnd_pct:.4f}"])
    sheet_assump.append(["Other income % of revenue", f"{assumptions.other_income_pct:.4f}"])
    sheet_assump.append(["Capex % of revenue", f"{assumptions.capex_pct:.4f}"])

    # Projection sheet
    sheet_proj = wb.create_sheet(title="Projections")
    header = ["Line Item"] + years_list
    sheet_proj.append(header)
    for label, series in projections.items():
        row = [label] + [series.get(y, "") for y in years_list]
        sheet_proj.append(row)

    wb.save(output_path)


def write_template_with_projections(
    baseline_path: Path,
    output_path: Path,
    assumptions: Assumptions,
    projections: Mapping[str, Mapping[int, Number]],
    note: str | None = None,
    ratios: Mapping[str, Number] | None = None,
) -> None:
    """
    Load the baseline workbook, fill assumptions/margins/projections in-place,
    and save to a new file so the layout matches the original template.
    """
    grid = load_grid_from_xlsx(baseline_path)
    year_row_idx, year_map = _find_year_row(grid)

    wb = load_workbook(baseline_path)
    ws = wb.active

    def _set_series(label: str, series: Mapping[int, Number]) -> None:
        row_idx = _find_row_index_by_label(grid, label)
        for col, year in year_map.items():
            if year in series:
                cell = ws[f"{col}{row_idx}"]
                # Skip merged header cells that cannot be assigned
                if cell.__class__.__name__ == "MergedCell":
                    continue
                cell.value = series[year]

    # Core line items
    for label, series in projections.items():
        try:
            _set_series(label, series)
        except KeyError:
            continue

    revenue = projections["Revenue"]
    cogs = projections["COGS"]
    gp = projections["Gross Profit"]
    sgna = projections["SG&A"]
    rnd = projections["R&D"]
    organic_ebitda = projections["Organic EBITDA"]
    total_ebitda = projections["Total EBITDA"]

    years_sorted = sorted(year_map.values())

    # Growth row under Revenue
    growth_row = _find_row_index_by_label(grid, "% growth")
    for col, year in year_map.items():
        prev_year = year - 1
        if year in revenue and prev_year in revenue:
            ws[f"{col}{growth_row}"] = (revenue[year] / revenue[prev_year]) - 1

    # Margins
    def _set_margin(label: str, numerator: Mapping[int, Number]) -> None:
        row_idx = _find_row_index_by_label(grid, label)
        for col, year in year_map.items():
            if year in numerator and year in revenue and revenue[year] != 0:
                ws[f"{col}{row_idx}"] = numerator[year] / revenue[year]

    _set_margin("% expenses", cogs)
    _set_margin("% margin", gp)

    # SG&A margin (row after SG&A)
    sgna_margin_row = _find_row_index_by_label(grid, "   % margin", label_col="C")
    # There are multiple "% margin" labels; locate the one immediately after SG&A row
    sgna_row = _find_row_index_by_label(grid, "SG&A")
    for row_idx in sorted(grid):
        if row_idx > sgna_row and _norm_label(grid[row_idx].get("C", "")) == _norm_label("% margin"):
            sgna_margin_row = row_idx
            break
    for col, year in year_map.items():
        if year in sgna and year in revenue and revenue[year] != 0:
            ws[f"{col}{sgna_margin_row}"] = sgna[year] / revenue[year]

    # R&D margin
    rnd_margin_row = _find_row_index_by_label(grid, "   % margin", label_col="C")
    rnd_row = _find_row_index_by_label(grid, "R&D")
    for row_idx in sorted(grid):
        if row_idx > rnd_row and _norm_label(grid[row_idx].get("C", "")) == _norm_label("% margin"):
            rnd_margin_row = row_idx
            break
    for col, year in year_map.items():
        if year in rnd and year in revenue and revenue[year] != 0:
            ws[f"{col}{rnd_margin_row}"] = rnd[year] / revenue[year]

    # EBITDA margin (organic / total share same %)
    ebitda_row = _find_row_index_by_label(grid, "Total EBITDA")
    ebitda_margin_row = None
    for row_idx in sorted(grid):
        if row_idx > ebitda_row and _norm_label(grid[row_idx].get("C", "")) == _norm_label("% margin"):
            ebitda_margin_row = row_idx
            break
    if ebitda_margin_row:
        for col, year in year_map.items():
            if year in total_ebitda and year in revenue and revenue[year] != 0:
                ws[f"{col}{ebitda_margin_row}"] = total_ebitda[year] / revenue[year]

    # Capex % revenue
    capex_pct_row = _find_row_index_by_label(grid, "% Revenue")
    for col, year in year_map.items():
        if year in projections["Capex"] and year in revenue and revenue[year] != 0:
            ws[f"{col}{capex_pct_row}"] = projections["Capex"][year] / revenue[year]

    # CAGR cell
    ws[f"N{year_row_idx}"] = assumptions.revenue_growth_cagr

    # Assumptions block in column Q (values to the right of the labels in column P)
    def _set_assumption(label: str, value: Number) -> None:
        row_idx = _find_row_index_by_label(grid, label, label_col="P")
        cell = ws[f"Q{row_idx}"]
        if cell.__class__.__name__ != "MergedCell":
            cell.value = value

    # Fill known template assumption slots if present
    try:
        _set_assumption("Revenue growth ", assumptions.revenue_growth_cagr)
        _set_assumption("COGS", assumptions.cogs_pct)
        _set_assumption("SG&A", assumptions.sgna_pct)
        _set_assumption("R&D", assumptions.rnd_pct)
        _set_assumption("Other Income", assumptions.other_income_pct)
        _set_assumption("CAPEX", assumptions.capex_pct)
    except Exception:
        pass

    # Dynamically write any additional ratios to the assumptions block in columns P/Q
    if ratios:
        existing_rows = [r for r in grid if _norm_label(str(grid[r].get("P", "")))]
        start_row = (max(existing_rows) + 1) if existing_rows else 2
        for label, value in ratios.items():
            ws[f"P{start_row}"] = label
            ws[f"Q{start_row}"] = value
            start_row += 1

    # Optional note about source/model in an unused corner
    if note:
        ws["V1"] = note

    wb.save(output_path)
