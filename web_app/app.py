from flask import Flask, request, redirect, url_for, send_file, render_template, flash, Response, session
import os
import sys
import subprocess
import uuid
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
from dotenv import load_dotenv
import requests
import sqlite3
import secrets
import smtplib
from email.message import EmailMessage
import json

# Load environment variables from .env if present
load_dotenv()

# Configuration
BASE_DIR = os.path.abspath(os.path.dirname(__file__)) + os.sep + ".."
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), 'outputs')
DATABASE_PATH = os.path.join(os.path.dirname(__file__), "users.db")
ALLOWED_EXTENSIONS = {'.html', '.htm'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Toggle LLM-based label mapping (Ollama). Default via env.
USE_LLM_DEFAULT = os.environ.get('USE_LLM_DEFAULT', '0').lower() in {'1', 'true', 'yes', 'on'}

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB uploads
app.secret_key = os.environ.get('WEB_APP_SECRET', 'dev-secret')
def allowed_file(filename):
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_EXTENSIONS


def get_db():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT UNIQUE NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            used INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            company_name TEXT,
            source_type TEXT,
            url_used TEXT,
            base_growth REAL,
            combined_score INTEGER,
            growth_bump REAL,
            external_used INTEGER,
            ai_sentiment_label TEXT,
            ai_sentiment_note TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )
    conn.commit()
    conn.close()

# Initialize DB once at startup
with app.app_context():
    init_db()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

# Alias for any legacy decorator usage
requires_auth = login_required


def run_pipeline():
    # Accept either a direct URL or a posted HTML file.
    url = request.form.get('url', '').strip()
    sec_id = request.form.get('sec_id', '').strip()
    file = request.files.get('file')
    company = request.form.get('company', '').strip()
    use_llm_flag_raw = request.form.get('use_llm_flag')
    use_external_news_raw = request.form.get("use_external_news")
    if use_llm_flag_raw is None:
        use_llm = True  # safe default: always use OpenAI mapping
    else:
        use_llm = (use_llm_flag_raw == '1')
    # Debug incoming form payload
    try:
        app.logger.info("request.form: %s", dict(request.form))
    except Exception:
        pass

    if not company:
        return None, None, None, "Please provide a company name"

    unique_suffix = datetime.utcnow().strftime('%Y%m%d%H%M%S') + '_' + uuid.uuid4().hex[:8]

    saved_path = None
    if url:
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (automated)'}
            resp = requests.get(url, headers=headers, timeout=20)
            resp.raise_for_status()
        except Exception as e:
            return None, None, None, f'Failed to fetch URL: {e}'

        saved_name = f"url_{unique_suffix}.html"
        saved_path = os.path.join(UPLOAD_FOLDER, saved_name)
        try:
            with open(saved_path, 'wb') as fh:
                fh.write(resp.content)
        except Exception as e:
            return None, None, None, f'Failed to save fetched HTML: {e}'
    elif file and file.filename:
        if not allowed_file(file.filename):
            return None, None, None, 'Unsupported file type. Upload HTML file.'
        filename = secure_filename(file.filename)
        saved_name = f"{os.path.splitext(filename)[0]}_{unique_suffix}{os.path.splitext(filename)[1]}"
        saved_path = os.path.join(UPLOAD_FOLDER, saved_name)
        file.save(saved_path)
    else:
        # No URL/file; allow if SEC id was provided
        if not sec_id:
            return None, None, None, 'No file selected, no URL provided, and no SEC identifier provided.'

    # create output paths
    mid_name = f"mid_product_{unique_suffix}.xlsx"
    safe_company = secure_filename(company) or "output"
    final_download_name = f"{safe_company}_financial_modeling.xlsx"
    final_name = f"final_{unique_suffix}.xlsx"
    mid_path = os.path.join(OUTPUT_FOLDER, mid_name)
    final_path = os.path.join(OUTPUT_FOLDER, final_name)
    # Build environment for subprocess with per-run external news flag
    env = os.environ.copy()
    env_default_ext = env.get("USE_EXTERNAL_OUTLOOK", "0").lower() in {"1", "true", "yes", "on"}
    has_ui_flag = use_external_news_raw is not None
    ui_flag = use_external_news_raw == "1"
    if has_ui_flag:
        use_external_outlook = ui_flag  # per-run override
    else:
        use_external_outlook = env_default_ext
    env["USE_EXTERNAL_OUTLOOK"] = "1" if use_external_outlook else "0"
    try:
        app.logger.info(
            "use_external_outlook derived: %s (raw=%s, env_default=%s)",
            use_external_outlook, use_external_news_raw, env_default_ext
        )
    except Exception:
        pass

    # Build command: use same Python interpreter running the app (respects venv)
    python_exec = sys.executable
    script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'unified_pipeline.py'))

    cmd = [python_exec, script_path, '--company', company, '--mid-product', mid_path, '--final', final_path]
    if saved_path:
        cmd.extend(['--html', saved_path])
    if sec_id:
        cmd.extend(['--sec-id', sec_id])
    if use_llm:
        cmd.append('--use-llm')

    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, check=False, env=env)
    except Exception as e:
        return None, None, None, f'Failed to run pipeline: {e}'

    stdout = proc.stdout
    stderr = proc.stderr
    returncode = proc.returncode
    if use_llm and "LLM mapping failed" in stdout + stderr:
        flash("LLM was requested but unavailable; fell back to non-LLM mapping.")

    if returncode != 0:
        return None, stdout, stderr, f"Pipeline failed (exit {returncode}). See logs below."

    if not os.path.exists(final_path) or os.path.getsize(final_path) == 0:
        return None, stdout, stderr, "Final file not found after processing."

    if sec_id and not saved_path:
        source_type = "sec_api"
    elif sec_id and saved_path:
        source_type = "sec_api+html"
    else:
        source_type = "url" if url else "file"

    return {
        "final_path": final_path,
        "final_download_name": final_download_name,
        "company": company,
        "url_used": url if url else None,
        "source_type": source_type
    }, stdout, stderr, None


@app.route('/', methods=['GET'])
@login_required
def home():
    return render_template('home.html')


@app.route('/app', methods=['GET'])
@login_required
def app_page():
    use_external_default = os.environ.get("USE_EXTERNAL_OUTLOOK", "0").lower() in {"1", "true", "yes", "on"}
    return render_template('app.html', use_llm_default=USE_LLM_DEFAULT, use_external_default=use_external_default)


@app.route('/help', methods=['GET'])
@login_required
def help_page():
    return render_template('help.html')


@app.route('/settings', methods=['GET'])
@login_required
def settings_page():
    return render_template('settings.html', use_llm_default=USE_LLM_DEFAULT)


@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    runs = []
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT created_at, company_name, source_type, url_used, base_growth, combined_score, growth_bump, external_used, ai_sentiment_label FROM runs WHERE user_id = ? ORDER BY created_at DESC LIMIT 50",
            (session.get("user_id"),)
        )
        runs = cur.fetchall()
        conn.close()
    except Exception as e:
        print(f"[WARN] Could not load runs: {e}")
    return render_template('dashboard.html', runs=runs)


@app.route('/generate', methods=['POST'])
@login_required
def generate():
    result, stdout, stderr, err = run_pipeline()
    if err:
        if request.form.get('ajax') == '1':
            body = f"{err}\n\nSTDOUT:\n{stdout}\n\nSTDERR:\n{stderr}"
            return Response(body, status=500, mimetype='text/plain')
        flash(err)
        return render_template('result.html', stdout=stdout, stderr=stderr, final_filename=None)

    final_path = result["final_path"]
    final_download_name = result["final_download_name"]

    # Try to log run
    try:
        # Pull sentiment metadata if available
        sentiment_file = os.path.abspath(os.path.join(BASE_DIR, "automodel/data/interim/last_sentiment.json"))
        combined_score = None
        bump_pct = None
        external_used = 0
        external_requested = 0
        ai_label = None
        ai_note = None
        if os.path.exists(sentiment_file):
            with open(sentiment_file, "r") as fh:
                sent = json.load(fh)
            raw_sent = sent.get("raw", sent)
            if isinstance(raw_sent, dict):
                combined_score = raw_sent.get("combined_score") if raw_sent.get("combined_score") is not None else raw_sent.get("score")
                ai_label = raw_sent.get("combined_label") or raw_sent.get("label")
                ev = raw_sent.get("summary") or raw_sent.get("justification_short")
                ai_note = ev
            external_used = 1 if sent.get("external_used") else 0
            external_requested = 1 if sent.get("external_requested") else 0
            # bump maybe in AI sheet but store from sentiment result mapping if present
        # attempt to read base growth from final workbook (Q5)
        base_growth = None
        try:
            import openpyxl
            wb = openpyxl.load_workbook(final_path, data_only=True)
            ws = wb.active
            base_growth = ws["Q5"].value
        except Exception:
            base_growth = None
        growth_bump = None
        try:
            # we wrote bump pct to AI_Sentiment sheet; attempt read
            import openpyxl
            wb2 = openpyxl.load_workbook(final_path, data_only=True)
            if "AI_Sentiment" in wb2.sheetnames:
                s = wb2["AI_Sentiment"]
                growth_bump = s["B4"].value
        except Exception:
            growth_bump = None
        try:
            app.logger.info(
                "sentiment summary: external_requested=%s external_used=%s ai_note=%s",
                external_requested, external_used, ai_note
            )
        except Exception:
            pass
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO runs (user_id, company_name, source_type, url_used, base_growth, combined_score, growth_bump, external_used, ai_sentiment_label, ai_sentiment_note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session.get("user_id"),
                result.get("company"),
                result.get("source_type"),
                result.get("url_used"),
                base_growth,
                combined_score,
                growth_bump,
                external_used,
                ai_label,
                (ai_note or "")[:500]
            )
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARN] Could not log run: {e}")

    if request.form.get('ajax') == '1':
        return send_file(final_path, as_attachment=True, download_name=final_download_name)

    return render_template('result.html', stdout=stdout, stderr=stderr, final_filename=final_download_name)


@app.route('/download/<filename>')
@login_required
def download(filename):
    fpath = os.path.join(OUTPUT_FOLDER, filename)
    if not os.path.exists(fpath):
        flash('File not found')
        return redirect(url_for('app_page'))
    return send_file(fpath, as_attachment=True, download_name=filename)


def send_reset_email(email: str, reset_link: str):
    server = os.environ.get("MAIL_SERVER")
    port = os.environ.get("MAIL_PORT")
    user = os.environ.get("MAIL_USERNAME")
    pwd = os.environ.get("MAIL_PASSWORD")
    sender = os.environ.get("MAIL_SENDER", "no-reply@example.com")
    use_tls = os.environ.get("MAIL_USE_TLS", "0") == "1"
    if not server or not port or not user or not pwd:
        print(f"PASSWORD RESET LINK (email not configured): {reset_link}")
        return
    try:
        msg = EmailMessage()
        msg["Subject"] = "Password reset instructions"
        msg["From"] = sender
        msg["To"] = email
        msg.set_content(f"You requested a password reset. Click this link to set a new password:\n{reset_link}\nIf you did not request this, you can ignore this email.")
        with smtplib.SMTP(server, int(port), timeout=10) as smtp:
            if use_tls:
                smtp.starttls()
            smtp.login(user, pwd)
            smtp.send_message(msg)
    except Exception as e:
        print(f"[WARN] Could not send reset email, falling back to log: {e}")
        print(f"PASSWORD RESET LINK (email fallback): {reset_link}")


def cleanup_tokens():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM password_reset_tokens WHERE expires_at <= ? OR used = 1", (datetime.utcnow(),))
        conn.commit()
        conn.close()
    except Exception:
        pass


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, email, password_hash FROM users WHERE email = ?", (email,))
        row = cur.fetchone()
        conn.close()
        if row and check_password_hash(row["password_hash"], password):
            session["user_id"] = row["id"]
            session["email"] = row["email"]
            try:
                app.logger.info("User logged in: %s (id=%s)", email, row["id"])
            except Exception:
                pass
            return redirect(url_for('home'))
        try:
            app.logger.info("Login failed for email: %s", email)
        except Exception:
            pass
        flash("Invalid email/password")
    return render_template('login.html')


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not email or '@' not in email:
            flash("Please enter a valid email.")
            return render_template('signup.html')
        if password != confirm:
            flash("Passwords do not match.")
            return render_template('signup.html')
        pw_hash = generate_password_hash(password)
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("INSERT INTO users (email, password_hash) VALUES (?, ?)", (email, pw_hash))
            conn.commit()
            user_id = cur.lastrowid
            conn.close()
            try:
                app.logger.info("User signed up: %s (id=%s)", email, user_id)
            except Exception:
                pass
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash("That email is already registered.")
        except Exception:
            flash("Could not create account. Please try again.")
    return render_template('signup.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    cleanup_tokens()
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(hours=1)
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT id FROM users WHERE email = ?", (email,))
            row = cur.fetchone()
            if row:
                user_id = row["id"]
                cur.execute(
                    "INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (?, ?, ?)",
                    (user_id, token, expires_at)
                )
                conn.commit()
                reset_link = url_for('reset_password', token=token, _external=True)
                send_reset_email(email, reset_link)
            conn.close()
        except Exception as e:
            print(f"[WARN] Forgot-password flow issue: {e}")
        flash("If an account with that email exists, we have sent password reset instructions.")
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    cleanup_tokens()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT prt.id, prt.user_id, prt.expires_at, prt.used, u.email FROM password_reset_tokens prt JOIN users u ON prt.user_id = u.id WHERE prt.token = ?",
        (token,)
    )
    row = cur.fetchone()
    if not row:
        conn.close()
        return render_template('reset_password_invalid.html')
    expires_at = datetime.fromisoformat(row["expires_at"]) if isinstance(row["expires_at"], str) else row["expires_at"]
    if row["used"] or expires_at <= datetime.utcnow():
        conn.close()
        return render_template('reset_password_invalid.html')

    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not password or password != confirm:
            flash("Passwords must match and not be empty.")
            conn.close()
            return render_template('reset_password.html', token=token)
        pw_hash = generate_password_hash(password)
        try:
            cur.execute("UPDATE users SET password_hash = ? WHERE id = ?", (pw_hash, row["user_id"]))
            cur.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = ?", (row["id"],))
            cur.execute("UPDATE password_reset_tokens SET used = 1 WHERE user_id = ? AND id != ?", (row["user_id"], row["id"]))
            conn.commit()
            flash("Your password has been reset. You can now log in.")
            conn.close()
            return redirect(url_for('login'))
        except Exception as e:
            print(f"[WARN] Password reset failed: {e}")
            flash("Could not reset password. Please try again.")
            conn.close()
            return render_template('reset_password.html', token=token)

    conn.close()
    return render_template('reset_password.html', token=token)


if __name__ == '__main__':
    # Bind to host/port from environment for cloud deployment (Render/Heroku/Railway)
    port = int(os.environ.get('PORT', 8501))
    host = os.environ.get('HOST', '0.0.0.0')
    debug = os.environ.get('FLASK_DEBUG', '0') in ('1', 'true', 'True')
    app.run(host=host, port=port, debug=debug)
