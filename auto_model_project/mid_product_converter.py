"""
Convert automodel CSV output to finmod-compatible Excel template.

This script transforms the extracted financial data from automodel into a format
that finmod expects (the Mid-Product).
"""

from pathlib import Path
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import math

def csv_to_finmod_format(
    csv_path: Path,
    template_path: Path,
    output_path: Path,
    company_name: str = "Company"
) -> Path:
    """
    Convert automodel CSV output to finmod-compatible Excel format.
    
    Args:
        csv_path: Path to IS_tidy_mapped_best_llm.csv from automodel
        template_path: Path to Baseline IS.xlsx template
        output_path: Path to write the converted Mid-Product Excel
        company_name: Company name for display
    
    Returns:
        Path to created Mid-Product Excel file
    """
    
    # Read the mapped data from automodel
    df = pd.read_csv(csv_path)
    
    # Ensure we have the right columns
    if 'coa' not in df.columns or 'year' not in df.columns or 'value' not in df.columns:
        raise ValueError("CSV must contain 'coa', 'year', and 'value' columns")
    
    # Remove rows without COA mapping
    df = df[df['coa'].notna()].copy()
    
    # Aggregate by COA and year (in case of duplicates)
    df_pivot = df.groupby(['coa', 'year'])['value'].sum().reset_index()
    
    # Load the template workbook
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Update company name
    ws['D1'] = company_name
    
    # Get the year row and column mapping
    # The template has years in row 4, starting from column E (column 5)
    year_map = {}
    for col_idx in range(5, 15):  # E to N
        cell = ws.cell(row=4, column=col_idx)
        if cell.value and isinstance(cell.value, (int, float)):
            year = int(cell.value)
            year_map[year] = col_idx
    
    if not year_map:
        raise ValueError("Could not find year columns in template row 4")
    
    # Define mapping from extracted COA to template rows
    coa_to_template_label = {
        'Revenue': 'Revenue',
        'COGS': 'COGS',
        'Gross Profit': 'Gross Profit',
        'Sales & Marketing': 'SG&A',  # Map to SG&A if separate not available
        'General & Administrative': 'SG&A',
        'Research & Development': 'R&D',
        'Depreciation & Amortization': 'R&D',  # Approximate mapping
        'Share-Based Compensation': 'R&D',  # Approximate mapping
        'Operating Income (EBIT)': 'Organic EBITDA',  # Approximate
        'Interest Expense': 'Other Income',
        'Interest Income': 'Other Income',
        'Other Income (Expense)': 'Other Income',
        'Income Tax Expense': 'Other Income',
        'Income Before Taxes': 'Organic EBITDA',  # Approximate
        'Net Income': 'Total EBITDA',  # Approximate
    }
    
    # Find template row indices
    template_rows = {}
    target_labels = set(coa_to_template_label.values())
    
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=3)  # Column C
        if cell.value:
            label = str(cell.value).strip()
            for target in target_labels:
                if label.lower() == target.lower():
                    template_rows[target] = row_idx
                    break
    
    # Aggregate extracted data by template label
    aggregated = {}
    for _, row in df_pivot.iterrows():
        coa = row['coa']
        year = int(row['year'])
        value = row['value']
        
        # Map to template label
        template_label = coa_to_template_label.get(coa)
        if template_label:
            if template_label not in aggregated:
                aggregated[template_label] = {}
            
            # Sum if multiple COAs map to same template label
            if year not in aggregated[template_label]:
                aggregated[template_label][year] = 0
            aggregated[template_label][year] += value
    
    # Fill in the template with aggregated data
    for template_label, year_values in aggregated.items():
        if template_label not in template_rows:
            continue
        
        row_idx = template_rows[template_label]
        for year, value in year_values.items():
            if year not in year_map:
                continue
            col_idx = year_map[year]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Save the modified workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    return output_path


def main():
    """Example usage."""
    csv_path = Path("automodel/data/interim/IS_tidy_mapped_best_llm.csv")
    template_path = Path("final-project_finmod-main/Inputs_Historical/Baseline IS.xlsx")
    output_path = Path("Mid-Product.xlsx")
    
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")
    
    result = csv_to_finmod_format(csv_path, template_path, output_path, "Apple Inc.")
    print(f"✅ Created Mid-Product: {result}")


if __name__ == "__main__":
    main()
