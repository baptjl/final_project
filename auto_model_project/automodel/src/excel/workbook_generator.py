# automodel/src/excel/workbook_generator.py
"""
Generate professional Excel workbooks from extracted financial data.
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class WorkbookGenerator:
    """Generate Excel workbooks with formatted financial statements."""
    
    def __init__(self):
        self.wb = None
        
    def create_income_statement(self, csv_path: Path, output_path: Path, company_name: str = "Company"):
        """Create formatted Income Statement Excel workbook."""
        
        # Read data
        df = pd.read_csv(csv_path)
        df = df[df['coa'].notna()].copy()
        
        # Create workbook
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Income Statement"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"{company_name} - Income Statement"
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        # Get years and CoA structure
        years = sorted(df['year'].unique())
        coa_order = ["Revenue", "COGS", "Gross Profit", "Sales & Marketing", 
                     "Research & Development", "General & Administrative",
                     "Depreciation & Amortization", "Share-Based Compensation",
                     "Operating Income (EBIT)", "Interest Expense", "Interest Income",
                     "Other Income (Expense)", "Income Before Taxes", 
                     "Income Tax Expense", "Net Income"]
        
        # Headers
        row = 3
        ws[f'A{row}'] = "Line Item"
        for i, year in enumerate(years):
            col_letter = get_column_letter(i + 2)
            ws[f'{col_letter}{row}'] = str(int(year))
            ws[f'{col_letter}{row}'].fill = header_fill
            ws[f'{col_letter}{row}'].font = header_font
        
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        
        # Data rows
        row = 4
        for coa in coa_order:
            coa_data = df[df['coa'] == coa]
            if coa_data.empty:
                continue
                
            ws[f'A{row}'] = coa
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
            
            for i, year in enumerate(years):
                year_data = coa_data[coa_data['year'] == year]
                col_letter = get_column_letter(i + 2)
                
                if not year_data.empty:
                    value = year_data['value'].sum()
                    ws[f'{col_letter}{row}'] = value / 1e9  # Convert to billions
                    ws[f'{col_letter}{row}'].number_format = '$#,##0.0,,"B"'
                else:
                    ws[f'{col_letter}{row}'] = None
                    
                ws[f'{col_letter}{row}'].border = border
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for i in range(len(years)):
            col_letter = get_column_letter(i + 2)
            ws.column_dimensions[col_letter].width = 15
        
        # Save
        self.wb.save(output_path)
        return output_path
    
    def create_summary_sheet(self, csv_path: Path, output_path: Path, company_name: str = "Company"):
        """Create a summary sheet with data overview."""
        
        df = pd.read_csv(csv_path)
        df = df[df['coa'].notna()].copy()
        
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Summary"
        
        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        
        # Title
        ws['A1'] = f"{company_name} - Financial Summary"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Data
        row = 3
        ws['A' + str(row)] = "Metric"
        ws['B' + str(row)] = "2024"
        ws['C' + str(row)] = "2023"
        ws['D' + str(row)] = "2022"
        
        for cell in ['A' + str(row), 'B' + str(row), 'C' + str(row), 'D' + str(row)]:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
        
        # Add metrics
        metrics = {
            "Revenue": "Revenue",
            "COGS": "COGS",
            "R&D Expense": "Research & Development",
            "Operating Income": "Operating Income (EBIT)",
        }
        
        row = 4
        years = [2024, 2023, 2022]
        
        for metric_name, coa_name in metrics.items():
            ws[f'A{row}'] = metric_name
            
            for col_idx, year in enumerate(years):
                col_letter = get_column_letter(col_idx + 2)
                value = df[(df['coa'] == coa_name) & (df['year'] == year)]['value'].sum()
                
                if value != 0:
                    ws[f'{col_letter}{row}'] = value / 1e9
                    ws[f'{col_letter}{row}'].number_format = '$#,##0.0,,"B"'
                else:
                    ws[f'{col_letter}{row}'] = "-"
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        self.wb.save(output_path)
        return output_path


def main():
    """Generate Excel workbooks from extracted data."""
    
    csv_path = Path("automodel/data/interim/IS_tidy_mapped_best_llm.csv")
    output_dir = Path("automodel/data/processed")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    generator = WorkbookGenerator()
    
    # Create Income Statement
    is_path = output_dir / "Apple_Income_Statement.xlsx"
    generator.create_income_statement(csv_path, is_path, "Apple Inc.")
    print(f"✅ Created: {is_path}")
    
    # Create Summary
    summary_path = output_dir / "Apple_Summary.xlsx"
    generator.create_summary_sheet(csv_path, summary_path, "Apple Inc.")
    print(f"✅ Created: {summary_path}")


if __name__ == "__main__":
    main()
